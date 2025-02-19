from django.shortcuts import render

# Create your views here.
import pandas as pd
from io import BytesIO
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from ..models import RawData, Case, Patent, CasePatent, CaseDetails
from django.core.exceptions import ValidationError
from django.utils.dateparse import parse_date
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from django.db import transaction
from django.db.models import Q
from ..serializers import CaseSerializer
import openpyxl
from django.http import HttpResponse
from collections import Counter
from django.http import JsonResponse
from django.db.models import Count, F
from ..models import Case, Patent, PlaintiffDetails, DefendantDetails
from rest_framework.decorators import api_view
from collections import defaultdict
from rest_framework.permissions import AllowAny, IsAuthenticated
# from ..utils import normalize_name, are_names_similar

class CaseStatisticsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Use Django ORM aggregation to get case status counts efficiently
            case_status_counts = (
                Case.objects.values("case_status")
                .annotate(count=Count("case_status"))
            )

            # Normalize case statuses
            status_counts = {
                (item["case_status"].lower().strip() if item["case_status"] else "unknown"): item["count"]
                for item in case_status_counts
            }

            # Fetch top tech areas
            tech_areas = (
                Patent.objects.values("tech_category")
                .annotate(case_count=Count("cases"))
                .order_by("-case_count")
            )

            # Fetch top plaintiffs and their law firms
            plaintiffs = (
                CaseDetails.objects.values("plaintiff")
                .annotate(case_count=Count("case"))
                .order_by("-case_count")[:25]
            )
            plaw_firms = (
                PlaintiffDetails.objects.exclude(plaintiff_law_firm="nan")
                .values("plaintiff_law_firm")
                .annotate(case_count=Count("case"))
                .order_by("-case_count")[:25]
            )

            # Fetch top defendants and their law firms
            defendants = (
                CaseDetails.objects.values("defendant")
                .annotate(case_count=Count("case"))
                .order_by("-case_count")[:25]
            )
            dlaw_firms = (
                DefendantDetails.objects.exclude(defendant_law_firm="nan")
                .values("defendant_law_firm")
                .annotate(case_count=Count("case"))
                .order_by("-case_count")[:25]
            )

            # --- Overlapping Law Firms ---
            plaintiff_firms = (
                PlaintiffDetails.objects.exclude(plaintiff_law_firm="nan")
                .values("plaintiff_law_firm")
                .annotate(plaintiff_case_count=Count("case"))
            )
            defendant_firms = (
                DefendantDetails.objects.exclude(defendant_law_firm="nan")
                .values("defendant_law_firm")
                .annotate(defendant_case_count=Count("case"))
            )

            plaintiff_firm_dict = {firm["plaintiff_law_firm"]: firm["plaintiff_case_count"] for firm in plaintiff_firms}
            defendant_firm_dict = {firm["defendant_law_firm"]: firm["defendant_case_count"] for firm in defendant_firms}

            overlapping_law_firms = [
                {
                    "entity": firm,
                    "plaintiff_case_count": plaintiff_firm_dict[firm],
                    "defendant_case_count": defendant_firm_dict[firm],
                    "total_cases": plaintiff_firm_dict[firm] + defendant_firm_dict[firm],
                }
                for firm in plaintiff_firm_dict.keys() & defendant_firm_dict.keys()
            ]
            overlapping_law_firms.sort(key=lambda x: x["total_cases"], reverse=True)

            # --- Overlapping Plaintiffs and Defendants ---
            plaintiff_entities = (
                CaseDetails.objects.values("plaintiff")
                .annotate(plaintiff_case_count=Count("case"))
            )
            defendant_entities = (
                CaseDetails.objects.values("defendant")
                .annotate(defendant_case_count=Count("case"))
            )

            plaintiff_dict = {p["plaintiff"]: p["plaintiff_case_count"] for p in plaintiff_entities}
            defendant_dict = {d["defendant"]: d["defendant_case_count"] for d in defendant_entities}

            overlapping_parties = [
                {
                    "entity": entity,
                    "plaintiff_case_count": plaintiff_dict[entity],
                    "defendant_case_count": defendant_dict[entity],
                    "total_cases": plaintiff_dict[entity] + defendant_dict[entity],
                }
                for entity in plaintiff_dict.keys() & defendant_dict.keys()
            ]
            overlapping_parties.sort(key=lambda x: x["total_cases"], reverse=True)

            response_data = {
                "total_cases": Case.objects.count(),
                "case_status_counts": status_counts,
                "top_defendants": defendants,
                "top_defendant_law_firms": dlaw_firms,
                "top_plaintiffs": plaintiffs,
                "top_plaintiff_law_firms": plaw_firms,
                "tech_area": tech_areas,
                "overlapping_law_firms": overlapping_law_firms,
                "overlapping_parties": overlapping_parties,  # Added this new data
            }

            return Response({"data": response_data}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Error processing request: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
class PlaintiffTypeCountView(APIView):
    permission_classes=[IsAuthenticated]
    
    def process_patent(self,patent):
        """
        Process a single patent to check if it was transferred and return associated cases.
        """
        original_names = normalize_name(patent.original_assignee)
        current_names = normalize_name(patent.current_assignee)

        # If names are not similar, consider it a transfer
        if not are_names_similar(original_names, current_names):
            # Fetch case numbers linked to this patent
            return list(CasePatent.objects.filter(patent=patent).values_list('case__case_no', flat=True))
        return []
    def get(self, request):
        try:
            # Aggregate count of unique plaintiff_type_and_size values from CaseDetails
            plaintiff_counts = (
                CaseDetails.objects.values("plaintiff_type")
                .annotate(count=Count("plaintiff_type"))
            )

            # Normalize plaintiff types
            plaintiff_type_counts = {
                (item["plaintiff_type"].strip().lower() if item["plaintiff_type"] else "unknown"): item["count"]
                for item in plaintiff_counts
            }

            defendant_counts = CaseDetails.objects.values_list("defendent_type", flat=True)

            # Dictionary to store aggregated counts
            defendant_type_counts = defaultdict(int)

            for entry in defendant_counts:
                if entry:  # Check if value exists
                    types = [t.strip().lower() for t in entry.split(",")]  # Split and normalize
                    for t in types:
                        defendant_type_counts[t] += 1  # Count each type separately

            # print("step2 started")
            # transferred_cases = set()

            # # Fetch all patents in bulk
            # patents = list(Patent.objects.all())

            # # Use ThreadPoolExecutor for parallel processing
            # with ThreadPoolExecutor(max_workers=5) as executor:  # Adjust the worker count as needed
            #     results = executor.map(self.process_patent, patents)

            # # Collect results from all threads
            # for case_list in results:
            #     transferred_cases.update(case_list)

            response_data = {
                "plaintiff_type_counts": plaintiff_type_counts,
                "defendant_type_counts": dict(defendant_type_counts),
                # "transferredCases":len(list(transferred_cases))
            }

            return Response({"data": response_data}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"Error processing request: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class IndustryStats(APIView):
    permission_classes=[IsAuthenticated]
    def get(self,request):
        try:
        # Fetch unique cases per industry
            industry_case_count = (
                Patent.objects
                .values('industry')  # Get industry field
                .annotate(unique_case_count=Count('cases', distinct=True))  # Count unique cases
                .order_by('-unique_case_count')
            )
            
            # Convert query result to a dictionary
            result = {entry['industry']: entry['unique_case_count'] for entry in industry_case_count if entry['industry']}
            
            return Response({"data":result}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": f"Error processing request: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

# class CaseEntityListing(APIView):
    
#     def get(self,request):
#     # Create an Excel workbook
#         try:
#             wb = openpyxl.Workbook()

#             # Sheet 1: Cases (Case Numbers)
#             ws_cases = wb.active
#             ws_cases.title = "Cases"
#             ws_cases.append(["Case Number"])
#             for case_no in Case.objects.values_list('case_no', flat=True):
#                 ws_cases.append([case_no])

#             # Sheet 2: Patents (Patent No, Title, Case Count)
#             ws_patents = wb.create_sheet(title="Patents")
#             ws_patents.append(["Patent Number", "Patent Title", "Case Count"])
#             patents = Patent.objects.all()
#             for patent in patents:
#                 case_count = patent.cases.count()
#                 ws_patents.append([patent.patent_no, patent.patent_title, case_count])

#             # Sheet 3: Plaintiffs (Plaintiff Name, Case Count)
#             ws_plaintiffs = wb.create_sheet(title="Plaintiffs")
#             ws_plaintiffs.append(["Plaintiff Name", "Case Count"])
#             plaintiff_counts = Counter(PlaintiffDetails.objects.values_list('plaintiff', flat=True))
#             for plaintiff, count in plaintiff_counts.items():
#                 ws_plaintiffs.append([plaintiff, count])

#             # Sheet 4: Defendants (Defendant Name, Case Count)
#             ws_defendants = wb.create_sheet(title="Defendants")
#             ws_defendants.append(["Defendant Name", "Case Count"])
#             defendant_counts = Counter(DefendantDetails.objects.values_list('defendant', flat=True))
#             for defendant, count in defendant_counts.items():
#                 ws_defendants.append([defendant, count])

#             # Sheet 5: Defendant Law Firms (Law Firm, Case Count)
#             ws_def_law_firms = wb.create_sheet(title="Defendant Law Firms")
#             ws_def_law_firms.append(["Defendant Law Firm", "Case Count"])
#             def_law_firm_counts = Counter(DefendantDetails.objects.values_list('defendant_law_firm', flat=True))
#             for law_firm, count in def_law_firm_counts.items():
#                 ws_def_law_firms.append([law_firm, count])

#             # Sheet 6: Plaintiff Law Firms (Law Firm, Case Count)
#             ws_plaintiff_law_firms = wb.create_sheet(title="Plaintiff Law Firms")
#             ws_plaintiff_law_firms.append(["Plaintiff Law Firm", "Case Count"])
#             plaintiff_law_firm_counts = Counter(PlaintiffDetails.objects.values_list('plaintiff_law_firm', flat=True))
#             for law_firm, count in plaintiff_law_firm_counts.items():
#                 ws_plaintiff_law_firms.append([law_firm, count])

#             # Create response
#             response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#             response['Content-Disposition'] = 'attachment; filename="cases_report.xlsx"'

#             # Save workbook to response
#             wb.save(response)

#             return response

#         except Exception as e:
#             return Response(
#                 {"error": f"Error processing request: {str(e)}"},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
#             )
            

class CaseEntityListing(APIView):
    permission_classes=[AllowAny]
    def get(self, request):
        try:
            wb = openpyxl.Workbook()

            # Sheet 1: Cases
            ws_cases = wb.active
            ws_cases.title = "Cases"
            ws_cases.append(["Case Number"])
            for case_no in Case.objects.values_list('case_no', flat=True).iterator():
                ws_cases.append([case_no])

            # Sheet 2: Patents
            ws_patents = wb.create_sheet(title="Patents")
            ws_patents.append(["Patent Number", "Patent Title", "Case Count"])
            patents = Patent.objects.annotate(case_count=Count('cases'))
            for patent in patents:
                ws_patents.append([patent.patent_no, patent.patent_title, patent.case_count])

            # Sheet 3: Plaintiffs
            ws_plaintiffs = wb.create_sheet(title="Plaintiffs")
            ws_plaintiffs.append(["Plaintiff Name", "Case Count"])
            plaintiff_counts = PlaintiffDetails.objects.values('plaintiff').annotate(count=Count('plaintiff'))
            for entry in plaintiff_counts:
                ws_plaintiffs.append([entry['plaintiff'], entry['count']])

            # Sheet 4: Defendants
            ws_defendants = wb.create_sheet(title="Defendants")
            ws_defendants.append(["Defendant Name", "Case Count"])
            defendant_counts = DefendantDetails.objects.values('defendant').annotate(count=Count('defendant'))
            for entry in defendant_counts:
                ws_defendants.append([entry['defendant'], entry['count']])

            # Sheet 5: Defendant Law Firms
            ws_def_law_firms = wb.create_sheet(title="Defendant Law Firms")
            ws_def_law_firms.append(["Defendant Law Firm", "Case Count"])
            def_law_firm_counts = DefendantDetails.objects.exclude(defendant_law_firm=None).values('defendant_law_firm').annotate(count=Count('defendant_law_firm'))
            for entry in def_law_firm_counts:
                ws_def_law_firms.append([entry['defendant_law_firm'], entry['count']])

            # Sheet 6: Plaintiff Law Firms
            ws_plaintiff_law_firms = wb.create_sheet(title="Plaintiff Law Firms")
            ws_plaintiff_law_firms.append(["Plaintiff Law Firm", "Case Count"])
            plaintiff_law_firm_counts = PlaintiffDetails.objects.exclude(plaintiff_law_firm=None).values('plaintiff_law_firm').annotate(count=Count('plaintiff_law_firm'))
            for entry in plaintiff_law_firm_counts:
                ws_plaintiff_law_firms.append([entry['plaintiff_law_firm'], entry['count']])

            # Write to memory buffer
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="cases_report.xlsx"'
            return response

        except Exception as e:
            return Response(
                {"error": f"Error processing request: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



@api_view(['GET'])
def most_sued_tech_areas(request):
    """Returns the top 10 most sued tech categories based on case count."""
    tech_areas = Patent.objects.values('tech_category').annotate(case_count=Count('cases')).order_by('-case_count')[:10]
    return Response(tech_areas)

@api_view(['GET'])
def top_defendants(request):
    """Returns the top 25 defendants and their law firms based on case count."""
    defendants = DefendantDetails.objects.values('defendant').annotate(case_count=Count('case')).order_by('-case_count')[:25]
    law_firms = DefendantDetails.objects.values('defendant_law_firm').annotate(case_count=Count('case')).order_by('-case_count')[:25]
    
    return Response({
        "top_defendants": defendants,
        "top_defendant_law_firms": law_firms
    })

@api_view(['GET'])
def top_plaintiffs(request):
    """Returns the top 25 plaintiffs and their law firms based on case count."""
    plaintiffs = PlaintiffDetails.objects.values('plaintiff').annotate(case_count=Count('case')).order_by('-case_count')[:25]
    law_firms = PlaintiffDetails.objects.values('plaintiff_law_firm').annotate(case_count=Count('case')).order_by('-case_count')[:25]
    
    return Response({
        "top_plaintiffs": plaintiffs,
        "top_plaintiff_law_firms": law_firms
    })
