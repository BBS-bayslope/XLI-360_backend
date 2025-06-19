from django.shortcuts import render
from memory_profiler import profile
# Create your views here.
import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import JsonResponse
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from ..models import RawData, Case, Patent, CasePatent, CaseDetails, PlaintiffDetails, DefendantDetails
from django.core.exceptions import ValidationError
from django.utils.dateparse import parse_date
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from django.db import transaction
from django.db.models import Q
from ..serializers import CaseSerializer
from django.db.models.functions import Upper, Lower
from django.db import connection
from rest_framework.permissions import AllowAny, IsAuthenticated
import json
from django.db.models.functions import Upper, Lower

class FileUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def process_chunk(self, rows):
        records = []
        required_columns = [
            'Case Number', 'Case Name', 'Case Status', 'Court Names',
            'Litigation Venues & Judicial Authorities', 'Related/Originating Cases',
            'Cause of Action', 'Accused Product', 'Judge', 'Number of Infringed Claims',
            '3rd Party Funding Involved', 'Case Strength Level', 'Recent Action',
            'Winning Amount', 'Winning Party', 'Other Possible Infringer', 'List of Prior Art',
            'Patent No', 'Type of Patent', 'Title', 'Original Assignee',
            'Current Assignee', 'Single Patent or Family Involved?',
            'Standard Essential Patent', 'Semiconductor Patent', 'Tech Centre',
            'Art Unit', 'Acquired Patent or Organic patent?', 'Assignee Timeline',
            'Industry', 'Technology Keywords', 'Tech Category', 'Reason of Allowance',
            'Plaintiff/Petitioner', 'Number of Plaintiff/petitioners', 'Plaintiff Type', 'Plaintiff Size', 'Defendant',
            'Defendant Type', 'Defendant Size', 'Number of Defendants', 'Stage', 'Chances of Winning'
        ]

        for row in rows:
            try:
                missing_columns = [col for col in required_columns if col not in row]
                if missing_columns:
                    raise KeyError(f"Missing columns: {', '.join(missing_columns)}")

                is_valid = bool(row.get('Case Number')) and bool(row.get('Case Complaint Date')) and bool(row['Case Name']) and bool(row['Patent No']) and bool(row['Plaintiff/Petitioner']) and bool(row['Defendant'])
        
                
                complaint_date = pd.to_datetime(row.get('Case Complaint Date'), errors='coerce')
                case_closed_date = pd.to_datetime(row.get('Case Closed Date'), errors='coerce')
                issue_date = pd.to_datetime(row.get('Patent Issued Date'), errors='coerce')
                expiry_date = pd.to_datetime(row.get('Patent Expiry Date'), errors='coerce')

                complaint_date = complaint_date.strftime('%Y-%m-%d') if pd.notnull(complaint_date) else None
                case_closed_date = case_closed_date.strftime('%Y-%m-%d') if pd.notnull(case_closed_date) else None
                issue_date = issue_date.strftime('%Y-%m-%d') if pd.notnull(issue_date) else None
                expiry_date = expiry_date.strftime('%Y-%m-%d') if pd.notnull(expiry_date) else None

                raw_data_instance = RawData(
                    case_no=row.get('Case Number'),
                    complaint_date=complaint_date,
                    case_name=row['Case Name'],
                    case_status=row['Case Status'],
                    court_name=row['Court Names'],
                    litigation_venues=row['Litigation Venues & Judicial Authorities'],
                    related_cases=row.get('Related/Originating Cases', ''),
                    case_closed_date=case_closed_date,
                    cause_of_action=row['Cause of Action'],
                    accused_product=row['Accused Product'],
                    assigned_judge=row.get('Judge', None),
                    number_of_infringed_claims=row.get('Number of Infringed Claims', 0) or 0,
                    third_party_funding_involved=row['3rd Party Funding Involved'],
                    # type_of_infringement=row['Type of Infringement'],
                    case_strength_level=row['Case Strength Level'],
                    recent_action=row.get('Recent Action', None),
                    winning_amount=row.get('Winning Amount', None),
                    winning_party=row.get('Winning Party', None),
                    other_possible_infringer=row.get('Other Possible Infringer', None),
                    list_of_prior_art=row.get('List of Prior Art', None),
                    patent_no=row.get('Patent No', None),
                    patent_type=row.get('Type of Patent', None),
                    patent_title=row.get('Title', None),
                    original_assignee=row['Original Assignee'],
                    current_assignee=row.get('Current Assignee', None),
                    issue_date=issue_date,
                    expiry_date=expiry_date,
                    single_or_multiple=row['Single Patent or Family Involved?'],
                    standard_patent=row['Standard Essential Patent'],
                    semiconductor_patent=row['Semiconductor Patent'],
                    tech_center=row['Tech Centre'],
                    art_unit=row['Art Unit'],
                    acquisition_type=row['Acquired Patent or Organic patent?'],
                    # assignee_timeline=row.get('Assignee Timeline', None),
                    industry=row['Industry'],
                    technology_keywords=row['Technology Keywords'],
                    tech_category=row['Tech Category'],
                    reason_of_allowance=row['Reason of Allowance'],
                    plaintiff=row.get('Plaintiff/Petitioner', None),
                    plaintiff_type=row['Plaintiff Type'],
                    plaintiff_size=row['Plaintiff Size'],
                    plaintiff_count=row['Number of Plaintiff/petitioners'],
                    # plaintiff_law_firm=row.get('Plaintiff Law Firm Name', None),
                    # plaintiff_attorney_name=row.get('PA Name 1', None),
                    # plaintiff_contact=row.get('PA Phone 1', None),
                    # plaintiff_email=row.get('PA Email 1', None),
                    defendant=row.get('Defendant', None),
                    defendent_type=row['Defendant Type'],
                    defendent_size=row['Defendant Size'],
                    defendent_count=row['Number of Defendants'],
                    # defendant_law_firm=row.get('Defendant Law Firm Name', None),
                    # defendant_attorney_name=row.get('DA Name 1', None),
                    # defendant_phone=row.get('DA Phone 1', None),
                    # defendant_email=row.get('DA Email 1', None),
                    stage=row['Stage'],
                    is_valid=is_valid
                )
                records.append(raw_data_instance)
            except KeyError as e:
                print(f"KeyError: {str(e)}")  # Log missing columns
                return e  # Returning KeyError causes an issue in the `post` method
            except Exception as e:
                print(f"Unexpected Error: {str(e)}")
                return e
        return records

    def get(self,request):
        create_stored_procedure()
        # Call the stored procedure to process data
        # try:
        #     with connection.cursor() as cursor:
        #         cursor.execute("SELECT process_raw_data()")

        #     return Response({"message": "File uploaded and processed successfully"})
        return Response({"message":"Success"})
        # except Exception as e:
        #     return Response({'error': f'Error processing file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def post(self, request):
        file = request.FILES.get('file')
        if file:
            try:
                excel_data = pd.ExcelFile(file)
                df = excel_data.parse(sheet_name=0, dtype=str)
                # df = pd.read_excel(file, dtype=str)  # Read all as strings for easier handling
                rows = df.to_dict('records')
                print(f"Rows shape: ({len(rows)}, {len(df.columns)})")
                chunk_size = 5000
                chunks = [rows[i:i + chunk_size] for i in range(0, len(rows), chunk_size)]
                print("chunks created")
                records_to_create = []
                cnt=0
                with ThreadPoolExecutor(max_workers=5) as executor:
                    futures = [executor.submit(self.process_chunk, chunk) for chunk in chunks]
                    for future in as_completed(futures):
                        cnt=cnt+1
                        records_to_create.extend(future.result())
                        # print("see ",cnt)

                with transaction.atomic():
                    RawData.objects.bulk_create(records_to_create, batch_size=5000)

                
                # Call the stored procedure to process data
                with connection.cursor() as cursor:
                    cursor.execute("SELECT process_raw_data()")

                # return JsonResponse({"message": "File uploaded and processed successfully"})
                return Response({'message': 'File processed and data inserted successfully!'}, status=status.HTTP_200_OK)
            except Exception as e:
                return Response({'error': f'Error processing file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)


def create_stored_procedure():
    print("hello")
    procedure_sql = """
    CREATE OR REPLACE FUNCTION process_raw_data()
    RETURNS VOID AS $$
    BEGIN
        -- Insert new cases from unprocessed raw data
        INSERT INTO mainapp_case (case_no, complaint_date, case_name, court_name, case_status, litigation_venues)
        SELECT DISTINCT case_no, complaint_date, case_name, court_name, case_status, litigation_venues
        FROM mainapp_rawdata
        WHERE is_processed = FALSE AND is_valid = TRUE
        ON CONFLICT (case_no) DO NOTHING;
        
        -- Insert or update case details
        INSERT INTO mainapp_casedetails (
            case_id, related_cases, case_closed_date, cause_of_action, accused_product,
            judge, number_of_infringed_claims, third_party_funding_involved,
            case_strength_level, recent_action, winning_amount, winning_party, other_possible_infringer,
            list_of_prior_art, plaintiff, defendant, plaintiff_type, plaintiff_size, plaintiff_count, defendent_type, defendent_size, defendent_count, chances_of_winning
        )
        SELECT DISTINCT ON (c.id)  -- Ensures one record per case_id
            c.id, rd.related_cases, rd.case_closed_date, rd.cause_of_action, rd.accused_product,
            rd.assigned_judge, rd.number_of_infringed_claims, rd.third_party_funding_involved,
            rd.case_strength_level, rd.recent_action, rd.winning_amount,
            rd.winning_party, rd.other_possible_infringer, rd.list_of_prior_art, rd.plaintiff, rd.defendant, rd.plaintiff_type, rd.plaintiff_size,
            rd.plaintiff_count, rd.defendent_type, rd.defendent_size, rd.defendent_count, rd.chances_of_winning
        FROM mainapp_rawdata rd
        JOIN mainapp_case c ON rd.case_no = c.case_no
        WHERE rd.is_processed = FALSE AND rd.is_valid = TRUE
        ON CONFLICT (case_id) DO UPDATE SET
            related_cases = EXCLUDED.related_cases,
            case_closed_date = EXCLUDED.case_closed_date,
            cause_of_action = EXCLUDED.cause_of_action,
            accused_product = EXCLUDED.accused_product,
            judge = EXCLUDED.judge,
            number_of_infringed_claims = EXCLUDED.number_of_infringed_claims,
            third_party_funding_involved = EXCLUDED.third_party_funding_involved,
            case_strength_level = EXCLUDED.case_strength_level,
            recent_action = EXCLUDED.recent_action,
            winning_amount = EXCLUDED.winning_amount,
            winning_party = EXCLUDED.winning_party,
            other_possible_infringer = EXCLUDED.other_possible_infringer,
            list_of_prior_art = EXCLUDED.list_of_prior_art,
            plaintiff = EXCLUDED.plaintiff,
            defendant = EXCLUDED.defendant,
            plaintiff_type = EXCLUDED.plaintiff_type,
            plaintiff_size = EXCLUDED.plaintiff_size,
            plaintiff_count = EXCLUDED.plaintiff_count,
            defendent_type = EXCLUDED.defendent_type,
            defendent_size = EXCLUDED.defendent_size,
            defendent_count = EXCLUDED.defendent_count,
            chances_of_winning = EXCLUDED.chances_of_winning;
            
        -- Insert new patents (ignoring duplicates)
        INSERT INTO mainapp_patent (
            patent_no, patent_type, patent_title, original_assignee, current_assignee, issue_date, expiry_date,
            single_or_multiple, standard_patent, semiconductor_patent, tech_center, art_unit, acquisition_type,
            assignee_timeline, industry, technology_keywords, tech_category, reason_of_allowance
        )
        SELECT DISTINCT 
            patent_no, patent_type, patent_title, original_assignee, current_assignee, issue_date, expiry_date,
            single_or_multiple, standard_patent, semiconductor_patent, tech_center, art_unit, acquisition_type,
            assignee_timeline, industry, technology_keywords, tech_category, reason_of_allowance
        FROM mainapp_rawdata
        WHERE 
            patent_no IS NOT NULL 
            AND is_processed = FALSE 
            AND is_valid = TRUE
        ON CONFLICT (patent_no) DO NOTHING;

        -- Insert case-patent mappings
        INSERT INTO mainapp_casepatent (case_id, patent_id, date_added)
        SELECT 
            c.id AS case_id,
            p.id AS patent_id,
            CURRENT_DATE  -- Adds today's date
        FROM mainapp_rawdata rd
        JOIN mainapp_case c ON rd.case_no = c.case_no
        JOIN mainapp_patent p ON rd.patent_no = p.patent_no
        WHERE rd.is_processed = FALSE AND rd.is_valid = TRUE
        ON CONFLICT DO NOTHING;


        -- Mark processed records
        UPDATE mainapp_rawdata SET is_processed = TRUE WHERE is_processed = FALSE AND is_valid = TRUE;
    END;
    $$ LANGUAGE plpgsql;
    """

    try:
        with connection.cursor() as cursor:
            cursor.execute(procedure_sql)
        print("Stored Procedure 'process_raw_data' created successfully.")
    except Exception as e:
        print(f"SQL Execution Error: {str(e)}")

# class FileUploadViewNew(APIView):
#     parser_classes = (MultiPartParser, FormParser)
#     permission_classes = [AllowAny]
#     def process_chunk(self, rows):
#         records = []
#         required_columns = [
#             'Case Number', 'Case Name', 'Case Status', 'Court Names',
#             'Litigation Venues & Judicial Authorities', 'Related/Originating Cases',
#             'Cause of Action', 'Accused Product', 'Judge', 'Number of Infringed Claims',
#             '3rd Party Funding Involved', 'Case Strength Level', 'Recent Action',
#             'Winning Amount', 'Winning Party', 'Other Possible Infringer', 'List of Prior Art',
#             'Patent No', 'Type of Patent', 'Title', 'Original Assignee',
#             'Current Assignee', 'Single Patent or Family Involved?',
#             'Standard Essential Patent', 'Semiconductor Patent', 'Tech Centre',
#             'Art Unit', 'Acquired Patent or Organic patent?', 'Assignee Timeline',
#             'Industry', 'Technology Keywords', 'Tech Category', 'Reason of Allowance',
#             'Plaintiff/Petitioner', 'Number of Plaintiff/petitioners', 'Plaintiff Type', 'Plaintiff Size', 'Defendant',
#             'Defendant Type', 'Defendant Size', 'Number of Defendants', 'Stage', 'Chances of Winning'
#         ]

#         for row in rows:
#             try:
#                 missing_columns = [col for col in required_columns if col not in row]
#                 if missing_columns:
#                     raise KeyError(f"Missing columns: {', '.join(missing_columns)}")

#                 is_valid = bool(row.get('Case Number')) and bool(row.get('Case Complaint Date')) and bool(row['Case Name']) and bool(row['Patent No']) and bool(row['Plaintiff/Petitioner']) and bool(row['Defendant'])

#                 complaint_date = pd.to_datetime(row.get('Case Complaint Date'), errors='coerce')
#                 case_closed_date = pd.to_datetime(row.get('Case Closed Date'), errors='coerce')
#                 issue_date = pd.to_datetime(row.get('Patent Issued Date'), errors='coerce')
#                 expiry_date = pd.to_datetime(row.get('Patent Expiry Date'), errors='coerce')

#                 complaint_date = complaint_date.strftime('%Y-%m-%d') if pd.notnull(complaint_date) else None
#                 case_closed_date = case_closed_date.strftime('%Y-%m-%d') if pd.notnull(case_closed_date) else None
#                 issue_date = issue_date.strftime('%Y-%m-%d') if pd.notnull(issue_date) else None
#                 expiry_date = expiry_date.strftime('%Y-%m-%d') if pd.notnull(expiry_date) else None

#                 raw_data_instance = RawData(
#                     case_no=row.get('Case Number','').upper(),
#                     complaint_date=complaint_date,
#                     case_name=row['Case Name'],
#                     case_status=row['Case Status'],
#                     court_name=row['Court Names'],
#                     litigation_venues=row['Litigation Venues & Judicial Authorities'],
#                     related_cases=row.get('Related/Originating Cases', ''),
#                     case_closed_date=case_closed_date,
#                     cause_of_action=row['Cause of Action'],
#                     accused_product=row['Accused Product'],
#                     assigned_judge=row.get('Judge', None),
#                     number_of_infringed_claims=row.get('Number of Infringed Claims', 0) or 0,
#                     third_party_funding_involved=row['3rd Party Funding Involved'],
#                     # type_of_infringement=row['Type of Infringement'],
#                     case_strength_level=row['Case Strength Level'],
#                     recent_action=row.get('Recent Action', None),
#                     winning_amount=row.get('Winning Amount', None),
#                     winning_party=row.get('Winning Party', None),
#                     other_possible_infringer=row.get('Other Possible Infringer', None),
#                     list_of_prior_art=row.get('List of Prior Art', None),
#                     patent_no=row.get('Patent No', '').upper(),
#                     patent_type=row.get('Type of Patent', None),
#                     patent_title=row.get('Title', None),
#                     original_assignee=row['Original Assignee'],
#                     current_assignee=row.get('Current Assignee', None),
#                     issue_date=issue_date,
#                     expiry_date=expiry_date,
#                     single_or_multiple=row['Single Patent or Family Involved?'],
#                     standard_patent=row['Standard Essential Patent'],
#                     semiconductor_patent=row['Semiconductor Patent'],
#                     tech_center=row['Tech Centre'],
#                     art_unit=row['Art Unit'],
#                     acquisition_type=row['Acquired Patent or Organic patent?'],
#                     # assignee_timeline=row.get('Assignee Timeline', None),
#                     industry=row['Industry'],
#                     technology_keywords=row['Technology Keywords'],
#                     tech_category=row['Tech Category'],
#                     reason_of_allowance=row['Reason of Allowance'],
#                     plaintiff=row.get('Plaintiff/Petitioner', None),
#                     plaintiff_type=row['Plaintiff Type'],
#                     plaintiff_size=row['Plaintiff Size'],
#                     plaintiff_count=row['Number of Plaintiff/petitioners'],
#                     # plaintiff_law_firm=row.get('Plaintiff Law Firm Name', None),
#                     # plaintiff_attorney_name=row.get('PA Name 1', None),
#                     # plaintiff_contact=row.get('PA Phone 1', None),
#                     # plaintiff_email=row.get('PA Email 1', None),
#                     defendant=row.get('Defendant', None),
#                     defendent_type=row['Defendant Type'],
#                     defendent_size=row['Defendant Size'],
#                     defendent_count=row['Number of Defendants'],
#                     # defendant_law_firm=row.get('Defendant Law Firm Name', None),
#                     # defendant_attorney_name=row.get('DA Name 1', None),
#                     # defendant_phone=row.get('DA Phone 1', None),
#                     # defendant_email=row.get('DA Email 1', None),
#                     stage=row['Stage'],
#                     chances_of_winning=row['Chances of Winning'],
#                     is_valid=is_valid
#                 )
#                 records.append(raw_data_instance)
#             except Exception as e:
#                 print(f"Error processing row: {str(e)}")
#         return records

#     def process_plaintiffs(self, plaintiffs_rows):
#         def clean_value(value):
#             """Returns an empty string for None, empty values, or 'NA'-like values."""
#             return "" if value in [None, "", "NA", "N/A", "na", "n/a", "-"] else value
#         plaintiffs=[]
#         for row in plaintiffs_rows:
#             try:
#                 if(row.get('Case Number')):
#                     case_instance = Case.objects.get(case_no=row.get('Case Number').upper())
#                 else:
#                     return Response({'error': 'Case Number is missing!'}, status=status.HTTP_400_BAD_REQUEST)
#                 plaintiffs.append(
#                     PlaintiffDetails(
#                         case=case_instance,
#                         plaintiff=clean_value(row.get('Plaintiff/Petitioner')),
#                         plaintiff_law_firm=clean_value(row.get("Plaintiff's Law Firm Name")),
#                         plaintiff_attorney_name=clean_value(row.get('PA Name 1')),
#                         plaintiff_contact=clean_value(row.get('PA Phone 1')),
#                         plaintiff_email=clean_value(row.get('PA Email 1'))
#                     ) )
#             except Case.DoesNotExist:
#                 print(f"Case with case_no {row.get('Case Number')} in plaintiff does not exist. Skipping.")
#         return plaintiffs

#     def process_defendants(self, defendants_rows):
#         def clean_value(value):
#             """Returns an empty string for None, empty values, or 'NA'-like values."""
#             return "" if value in [None, "", "NA", "N/A", "na", "n/a", "-"] else value

#         defendants = []
#         for row in defendants_rows:
#             try:
#                 if(row.get('Case Number')):
#                     case_instance = Case.objects.get(case_no=row.get('Case Number').upper())
#                 else:
#                     return Response({'error': 'Case Number is missing!'}, status=status.HTTP_400_BAD_REQUEST)
#                 defendants.append(
#                     DefendantDetails(
#                     case=case_instance,
#                     defendant=clean_value(row.get('Defendant')),
#                     defendant_law_firm=clean_value(row.get('Defendant Law Firm Name')),
#                     defendant_attorney_name=clean_value(row.get('DA Name 1')),
#                     defendant_phone=clean_value(row.get('DA Phone 1')),
#                     defendant_email=clean_value(row.get('DA Email 1'))
#                 )) 
#             except Case.DoesNotExist:
#                 print(f"Case with case_no {row.get('Case Number')} in defendant does not exist. Skipping.")
#         return defendants

#     def process_AssigneeChunk(self,chunk):
#         """
#         Processes a chunk of data and prepares it for bulk update.
#         """
#         patents_to_update = []

#         for row in chunk:
#             patent_no = str(row['Patent_No']).upper().strip()
#             total_assignments = row.get('Total_assignments', 0)   # handling empty row having no assignements detials assigning them zero
#             if pd.isna(total_assignments) or total_assignments == '':
#                 total_assignments = 0
#             else:
#                 total_assignments = int(total_assignments)  # Convert safely

#             assignments = []
#             for i in range(1, total_assignments + 1):  
#                 execution_date = row.get(f'Execution_Date_{i}', None)
#                 assignor = row.get(f'Assignors_{i}', None)
#                 assignee = row.get(f'Assignee_{i}', None)

#                 if pd.notna(execution_date) and pd.notna(assignor) and pd.notna(assignee):
#                     assignments.append({
#                         "execution_date": str(execution_date).strip(),
#                         "assignor": str(assignor).strip(),
#                         "assignee": str(assignee).strip()
#                     })

#             assignee_timeline_json = json.dumps({"assignments": assignments})

#             try:
#                 patent = Patent.objects.get(patent_no=patent_no)
#                 patent.assignee_timeline = assignee_timeline_json
#                 patents_to_update.append(patent)
#             except Patent.DoesNotExist:
#                 print(patent_no, "patent missing")
#                 pass  # Skip if patent not found

#         # Bulk update to optimize database write
#         # if patents_to_update:
#         #     Patent.objects.bulk_update(patents_to_update, ['assignee_timeline'])

#         return patents_to_update # Return count of updated records


#     def get(self,request):
#         create_stored_procedure()
#         # Call the stored procedure to process data
#         # try:
#         #     with connection.cursor() as cursor:
#         #         cursor.execute("SELECT process_raw_data()")

#         #     return Response({"message": "File uploaded and processed successfully"})
#         return Response({"message":"Success"})
#         # except Exception as e:
#         #     return Response({'error': f'Error processing file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

#     def post(self, request):
#         file = request.FILES.get('file')
#         if file:
#             try:
#                 excel_data = pd.ExcelFile(file)
#                 raw_data_df = excel_data.parse(sheet_name=0, dtype=str)
#                 # plaintiffs_df = excel_data.parse(sheet_name=1, dtype=str)
#                 # defendants_df = excel_data.parse(sheet_name=2, dtype=str)
#                 # assigneeTimeline_df=excel_data.parse(sheet_name=3, dtype=str)

#                 raw_data_rows = raw_data_df.to_dict('records')
#                 # plaintiffs_rows = plaintiffs_df.to_dict('records')
#                 # defendants_rows = defendants_df.to_dict('records')
#                 # assignee_rows=assigneeTimeline_df.to_dict('records')

#                 chunk_size = 500
#                 raw_data_chunks = [raw_data_rows[i:i + chunk_size] for i in range(0, len(raw_data_rows), chunk_size)]
#                 print("chunks created")
#                 with ThreadPoolExecutor(max_workers=4) as executor:
#                     raw_data_futures = [executor.submit(self.process_chunk, chunk) for chunk in raw_data_chunks]

#                     records_to_create = []
#                     for future in as_completed(raw_data_futures):
#                         records_to_create.extend(future.result())

#                 print("updating...")
#                 # Step 1: Insert RawData and trigger stored procedure to create Case records
#                 with transaction.atomic():
#                     RawData.objects.bulk_create(records_to_create, batch_size=500)
                
                
#                 with connection.cursor() as cursor:
#                     cursor.execute("SELECT process_raw_data()")  # Ensure Case records are created

#                 # Step 3: Process Plaintiffs & Defendants with the mapped Case records
#                 # plaintiffs_chunks = [plaintiffs_rows[i:i + chunk_size] for i in range(0, len(plaintiffs_rows), chunk_size)]
#                 # defendants_chunks = [defendants_rows[i:i + chunk_size] for i in range(0, len(defendants_rows), chunk_size)]
#                 # with ThreadPoolExecutor(max_workers=4) as executor:
#                 #     plaintiffs_futures = [executor.submit(self.process_plaintiffs, chunk) for chunk in plaintiffs_chunks]
#                 #     defendants_futures = [executor.submit(self.process_defendants, chunk) for chunk in defendants_chunks]

#                 #     plaintiffs_to_create = []
#                 #     defendants_to_create = []
#                 #     for future in as_completed(plaintiffs_futures):
#                 #         plaintiffs_to_create.extend(future.result())

#                 #     for future in as_completed(defendants_futures):
#                 #         defendants_to_create.extend(future.result())

#                 # # Step 4: Insert Plaintiffs & Defendants into DB
#                 # with transaction.atomic():
#                 #     PlaintiffDetails.objects.bulk_create(plaintiffs_to_create, batch_size=500)
#                 #     DefendantDetails.objects.bulk_create(defendants_to_create, batch_size=500)
                     
#                 # # print("2nd step completed")
#                 # timeline_chunks = [assignee_rows[i:i + chunk_size] for i in range(0, len(assignee_rows), chunk_size)]
#                 # with ThreadPoolExecutor(max_workers=4) as executor:
#                 #     assignee_futures = [executor.submit(self.process_AssigneeChunk, chunk) for chunk in timeline_chunks]
                    
#                 #     assignee_to_update = []
#                 #     for future in as_completed(assignee_futures):
#                 #         assignee_to_update.extend(future.result())
                        
#                 # with transaction.atomic():
#                 #     Patent.objects.bulk_update(assignee_to_update, ['assignee_timeline']) 
                    
                
#                 return Response({'message': 'File processed and data inserted successfully!'}, status=status.HTTP_200_OK)

#             except Exception as e:
#                 return Response({'error': f'Error processing file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#         return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)





# import pandas as pd
# import json
# from rest_framework.views import APIView
# from rest_framework.parsers import MultiPartParser, FormParser
# from rest_framework.permissions import AllowAny
# from rest_framework.response import Response
# from rest_framework import status
# from concurrent.futures import ThreadPoolExecutor, as_completed
# from django.db import transaction, connection
# from .models import RawData
import logging
import time
import openpyxl
from io import BytesIO

# Configure logging
logging.basicConfig(level=logging.INFO)


# from .sync_module import sync_raw_to_case

from mainapp.models import RawData, Case, Patent, CasePatent 
logger = logging.getLogger(__name__)

class FileUploadViewNew(APIView):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = []

    def safe_str(self, value, field_name):
        try:
            if value is None:
                return ''
            return str(value).upper()
        except AttributeError as e:
            logger.error(f"Failed to convert {field_name} to upper: {str(e)}. Value: {value}")
            return str(value) if value is not None else ''

    def process_chunk(self, rows):
        records = []
        required_columns = [
            'Case Number', 'Case Name', 'Case Status', 'Court Names',
            'Litigation Venues & Judicial Authorities', 'Related/Originating Cases',
            'Cause of Action', 'Accused Product', 'Judge', 'Number of Infringed Claims',
            '3rd Party Funding Involved', 'Case Strength Level', 'Recent Action',
            'Winning Amount', 'Winning Party', 'Other Possible Infringer', 'List of Prior Art',
            'Patent No', 'Type of Patent', 'Title', 'Original Assignee',
            'Current Assignee', 'Single Patent or Family Involved?',
            'Standard Essential Patent', 'Semiconductor Patent', 'Tech Centre',
            'Art Unit', 'Acquired Patent or Organic patent?', 'Assignee Timeline',
            'Industry', 'Technology Keywords', 'Tech Category', 'Reason of Allowance',
            'Plaintiff/Petitioner', 'Number of Plaintiff/petitioners', 'Plaintiff Type', 'Plaintiff Size', 'Defendant',
            'Stage', 'Chances of Winning', 'Case Complaint Date'
        ]

        for row in rows:
            try:
                if not all(row.get(col) for col in ['Case Number', 'Case Name', 'Patent No']):
                    logger.warning(f"Skipping row with missing key fields: {row}")
                    continue

                missing_columns = [col for col in required_columns if col not in row]
                if missing_columns:
                    raise KeyError(f"Missing columns: {', '.join(missing_columns)}")

                is_valid = all([
                    row.get('Case Number'), row.get('Case Complaint Date'),
                    row.get('Case Name'), row.get('Patent No'),
                    row.get('Plaintiff/Petitioner'), row.get('Defendant')
                ])

                # Format dates
                def parse_date(val): return pd.to_datetime(val, errors='coerce').strftime('%Y-%m-%d') if pd.notnull(
                    pd.to_datetime(val, errors='coerce')) else None

                raw_data_instance = RawData(
                    case_no=self.safe_str(row.get('Case Number', ''), 'Case Number'),
                    complaint_date=parse_date(row.get('Case Complaint Date')),
                    case_name=self.safe_str(row.get('Case Name', ''), 'Case Name'),
                    case_status=self.safe_str(row.get('Case Status', ''), 'Case Status'),
                    court_name=self.safe_str(row.get('Court Names', ''), 'Court Names'),
                    litigation_venues=self.safe_str(row.get('Litigation Venues & Judicial Authorities', ''), 'Litigation Venues'),
                    related_cases=str(row.get('Related/Originating Cases', '')),
                    case_closed_date=parse_date(row.get('Case Closed Date')),
                    cause_of_action=self.safe_str(row.get('Cause of Action', ''), 'Cause of Action'),
                    accused_product=self.safe_str(row.get('Accused Product', ''), 'Accused Product'),
                    assigned_judge=str(row.get('Judge', '')),
                    number_of_infringed_claims=int(row.get('Number of Infringed Claims') or 0),
                    third_party_funding_involved=self.safe_str(row.get('3rd Party Funding Involved', ''), 'Funding'),
                    case_strength_level=self.safe_str(row.get('Case Strength Level', ''), 'Strength'),
                    recent_action=str(row.get('Recent Action', '')),
                    winning_amount=str(row.get('Winning Amount', '')),
                    winning_party=str(row.get('Winning Party', '')),
                    other_possible_infringer=str(row.get('Other Possible Infringer', '')),
                    list_of_prior_art=str(row.get('List of Prior Art', '')),
                    patent_no=self.safe_str(row.get('Patent No', ''), 'Patent No'),
                    patent_type=str(row.get('Type of Patent', '')),
                    patent_title=str(row.get('Title', '')),
                    original_assignee=str(row.get('Original Assignee', '')),
                    current_assignee=str(row.get('Current Assignee', '')),
                    issue_date=parse_date(row.get('Patent Issued Date')),
                    expiry_date=parse_date(row.get('Patent Expiry Date')),
                    single_or_multiple=self.safe_str(row.get('Single Patent or Family Involved?', ''), 'Single/Family'),
                    standard_patent=self.safe_str(row.get('Standard Essential Patent', ''), 'Standard'),
                    semiconductor_patent=self.safe_str(row.get('Semiconductor Patent', ''), 'Semiconductor'),
                    tech_center=self.safe_str(row.get('Tech Centre', ''), 'Tech Centre'),
                    art_unit=self.safe_str(row.get('Art Unit', ''), 'Art Unit'),
                    acquisition_type=str(row.get('Acquired Patent or Organic patent?', '')),
                    industry=self.safe_str(row.get('Industry', ''), 'Industry'),
                    technology_keywords=str(row.get('Technology Keywords', '')),
                    tech_category=self.safe_str(row.get('Tech Category', ''), 'Tech Category'),
                    reason_of_allowance=str(row.get('Reason of Allowance', '')),
                    plaintiff=str(row.get('Plaintiff/Petitioner', '')),
                    plaintiff_type=self.safe_str(row.get('Plaintiff Type', ''), 'Plaintiff Type'),
                    plaintiff_size=self.safe_str(row.get('Plaintiff Size', ''), 'Plaintiff Size'),
                    defendant=str(row.get('Defendant', '')),
                    stage=str(row.get('Stage', '')),
                    chances_of_winning=str(row.get('Chances of Winning', '')),
                    is_valid=is_valid
                )
                records.append(raw_data_instance)
            except Exception as e:
                logger.error(f"Error processing row: {str(e)}. Row data: {row}")
        return records

    # def sync_raw_to_case(self):
    #     logger.info(f"Total RawData valid rows: {RawData.objects.filter(is_valid=True).count()}")
    #     logger.info(f"Total unsynced rows: {RawData.objects.filter(is_valid=True, is_synced=False).count()}")

    #     valid_raws = RawData.objects.filter(is_valid=True, is_synced=False)  
    #     for r in valid_raws:
    #         case_obj, _ = Case.objects.get_or_create(
    #             case_no=r.case_no,
    #             defaults={
    #                 'complaint_date': r.complaint_date,
    #                 'case_name': r.case_name,
    #                 'case_status': r.case_status,
    #                 'court_name': r.court_name,
    #                 'litigation_venues': r.litigation_venues,
    #             }
    #         )
    #         patent_obj, _ = Patent.objects.get_or_create(
    #             patent_no=r.patent_no,
    #             defaults={
    #                 'patent_title': r.patent_title,
    #                 'tech_category': r.tech_category
    #             }
    #         )
    #         CasePatent.objects.get_or_create(case=case_obj, patent=patent_obj)
    #         r.is_synced = True
    #         # r.save(update_fields=["is_synced"])
    #         # logger.info(f"Synced case {r.case_no} and patent {r.patent_no}")
    #         batch.append(r)

    # # ✅ Efficiently update all at once
    # RawData.objects.bulk_update(batch, ['is_synced'], batch_size=1000)
    # logger.info(f"Bulk updated {len(batch)} records as synced.")

# from django.db import transaction

    def sync_raw_to_case(self):
        logger.info("Syncing RawData → Case/Patent/CasePatent...")

        valid_raws = RawData.objects.filter(is_valid=True, is_synced=False)[:1000]
        logger.info(f"Total valid unsynced rows: {valid_raws.count()}")

        batch_to_update = []

        with transaction.atomic():
            for r in valid_raws.iterator(chunk_size=1000):  # Efficient memory usage
                try:
                    # Create or get Case
                    case_obj, _ = Case.objects.get_or_create(
                        case_no=r.case_no,
                        defaults={
                            'complaint_date': r.complaint_date,
                            'case_name': r.case_name,
                            'case_status': r.case_status,
                            'court_name': r.court_name,
                            'litigation_venues': r.litigation_venues,
                        }
                    )

                    # Create or get Patent
                    patent_obj, _ = Patent.objects.get_or_create(
                        patent_no=r.patent_no,
                        defaults={
                            'patent_title': r.patent_title,
                            'tech_category': r.tech_category
                        }
                    )

                    # Create CasePatent relation if not exists
                    CasePatent.objects.get_or_create(case=case_obj, patent=patent_obj)

                    r.is_synced = True
                    batch_to_update.append(r)

                    # Bulk update in chunks of 1000
                    if len(batch_to_update) >= 1000:
                        RawData.objects.bulk_update(batch_to_update, ['is_synced'])
                        logger.info(f"Updated {len(batch_to_update)} records as synced.")
                        batch_to_update.clear()

                except Exception as e:
                    logger.error(f"Error syncing RawData ID {r.id}: {str(e)}")

        # Update any remaining records
            if batch_to_update:
                RawData.objects.bulk_update(batch_to_update, ['is_synced'])
                logger.info(f"Final update: {len(batch_to_update)} records synced.")

        logger.info("RawData sync completed.")


    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            chunk_size = 200
            start_time = time.time()
            logger.info("Starting file processing")

            file_content = BytesIO(file.read())
            workbook = openpyxl.load_workbook(file_content, read_only=True, data_only=True)
            sheet_name = workbook.sheetnames[0]
            sheet = workbook[sheet_name]
            row_count = sum(1 for _ in sheet.iter_rows(min_row=2))
            # row_count = sheet.max_row - 1
            print("Total data rows:", row_count)

            def read_sheet_in_chunks(sheet_name, chunksize):
                sheet = workbook[sheet_name]
                headers = [cell.value for cell in next(sheet.iter_rows())]
                rows = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    rows.append(row_dict)
                    if len(rows) >= chunksize:
                        yield rows
                        rows = []
                if rows:
                    yield rows

            raw_data_records = []
            total_rows_processed = 0
            for chunk in read_sheet_in_chunks(sheet_name, chunk_size):
                with ThreadPoolExecutor(max_workers=1) as executor:
                    futures = [executor.submit(self.process_chunk, chunk)]
                    for future in as_completed(futures):
                        raw_data_records.extend(future.result())
                total_rows_processed += len(chunk)

            logger.info("Inserting into RawData table...")
            with transaction.atomic():
                RawData.objects.bulk_create(raw_data_records, batch_size=200)
                with connection.cursor() as cursor:
                    cursor.execute("SELECT process_raw_data()")  # optional stored proc

            logger.info("Syncing RawData → Case/Patent/CasePatent...")
            self.sync_raw_to_case()

            workbook.close()
            total_time = time.time() - start_time
            return Response({
                'message': f'Upload complete. {total_rows_processed} rows processed.',
                'duration': f"{total_time:.2f} seconds"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Upload failed: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



